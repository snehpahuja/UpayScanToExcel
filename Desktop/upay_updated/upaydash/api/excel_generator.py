"""
Excel file generator for different document categories
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import datetime

class ExcelGenerator:
    
    @staticmethod
    def generate_excel(document, extracted_data):
        """
        Generate Excel file based on document category
        
        Args:
            document: Document object
            extracted_data: QuerySet of ExtractedData
            
        Returns:
            BytesIO object containing Excel file
        """
        category = document.category.category_name
        
        if category == 'student_marksheet':
            return ExcelGenerator._generate_grades_excel(extracted_data)
        elif category == 'attendance_sheet':
            return ExcelGenerator._generate_attendance_excel(extracted_data)
        else:
            return ExcelGenerator._generate_generic_excel(extracted_data)
    
    @staticmethod
    def _generate_grades_excel(extracted_data):
        """Generate Excel for grade sheets"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Grades"
        
        # Headers
        headers = ['Student Name', 'Math', 'Science', 'English', 'History', 'Geography', 'Total', 'Average']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Group data by student
        students = {}
        for field in extracted_data:
            if 'name' in field.field_name:
                student_num = field.field_name.split('_')[1]
                students[student_num] = {'name': field.field_value, 'marks': {}}
            elif any(subj in field.field_name.lower() for subj in ['math', 'science', 'english', 'history', 'geography']):
                student_num = field.field_name.split('_')[1]
                subject = field.field_name.split('_')[-1]
                if student_num in students:
                    students[student_num]['marks'][subject] = int(field.field_value) if field.field_value.isdigit() else 0
        
        # Add data
        for student_num in sorted(students.keys(), key=int):
            student_data = students[student_num]
            marks = student_data['marks']
            
            row = [
                student_data['name'],
                marks.get('math', 0),
                marks.get('science', 0),
                marks.get('english', 0),
                marks.get('history', 0),
                marks.get('geography', 0),
                sum(marks.values()),
                round(sum(marks.values()) / len(marks), 2) if marks else 0
            ]
            ws.append(row)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file
    
    @staticmethod
    def _generate_attendance_excel(extracted_data):
        """Generate Excel for attendance sheets"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        
        # Headers
        headers = ['Student Name'] + [f'Day {i}' for i in range(1, 32)] + ['Present', 'Absent', 'Excused', '% Attendance']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data (simplified for mock)
        # In real implementation, parse extracted_data properly
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file
    
    @staticmethod
    def _generate_generic_excel(extracted_data):
        """Generate generic Excel for other categories"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Headers
        ws.append(['Field Name', 'Field Value', 'Confidence Score'])
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Add data
        for field in extracted_data:
            ws.append([field.field_name, field.field_value, f"{field.confidence_score}%"])
        
        # Adjust widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file